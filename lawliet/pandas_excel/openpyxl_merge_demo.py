# @Time    : 2020/12/3 8:13 下午
# @Author  : luoh
# @Email   : luohenghlx@163.com
# @File    : openpyxl_merge_demo.py
# @Software: PyCharm
# @Description:


import openpyxl



wb = openpyxl.load_workbook('pandas_excel_test.xlsx')

print(wb.sheetnames)

wb2 = openpyxl.load_workbook('pandas_excel_test_pyxl.xlsx')

print(wb2.sheetnames)


wb_new = openpyxl.Workbook()
st1 = wb_new.copy_worksheet(wb.active)
st2 = wb_new.copy_worksheet(wb2.active)

wb_new.save('merged.xlsx')
print('finish......')