# @Time    : 2020/11/12 6:58 下午
# @Author  : luoh
# @Email   : luohenghlx@163.com
# @File    : openpyxl_demo.py
# @Software: PyCharm
# @Description:


import pandas as pd
import numpy as np
from datetime import datetime, timedelta,time
import openpyxl
from openpyxl.styles import Font, colors, Alignment

## read excel
wb = openpyxl.load_workbook('pandas_excel_test.xlsx')
sht = wb.active
data = [121,124,135,615]
sht.append(data)
cell_range = sht['A2':'D3']
for a in cell_range:
    print(a)
    for aa in a:
        print(aa.value)
wb.save('pandas_excel_test_pyxl.xlsx')



## new create
wb2 = openpyxl.Workbook()
# 激活 worksheet
ws = wb2.active
ws.append(data)
sheet_2=wb2.create_sheet('new sheet')
sheet_2.append(['This is A1', 'This is B1', 'This is C1'])
sheet_2.append({'A' : 'This is A1', 'C' : 'This is C1'})
sheet_2.append({1 : 'This is 3 A1', 3 : 'This is 3 C1'})
## 改变 sheet 标签按钮颜色
sheet_2.sheet_properties.tabColor = "1072BA"

bold_itatic_24_font = Font(name='等线', size=12, italic=True, color='34123d', bold=True)

sheet_2['A1'].font = bold_itatic_24_font

#更改单元格内容的两个方法
sheet_2["E1"] = "这是我输入的"
sheet_2.cell(row=5, column=5, value="这也是我输入的")

#进入一个指定工作范围
cell_range = sheet_2['F5':'D10']
col_getc = sheet_2['C']
col_getcd = sheet_2['C:D']
row_get10 = sheet_2[10]
row_get5and10 = sheet_2[5:10]
wb2.save('pandas_excel_test_pyxl_new.xlsx')
print('finish......')